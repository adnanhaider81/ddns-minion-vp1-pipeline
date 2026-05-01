#!/usr/bin/env python3
from pathlib import Path
import re
import sys

ROOT = Path(__file__).resolve().parents[1]
REF_FASTA = ROOT / "references" / "references.all.unique.fasta"
PRIMERS_XLSX = ROOT / "resources" / "BarcodedPrimers.xlsx"
BARCODES_CSV = ROOT / "examples" / "barcodes.csv"


def fail(message: str) -> None:
    raise SystemExit(f"ERROR: {message}")


def read_fasta(path: Path):
    headers = []
    lengths = []
    name = None
    seq = []
    with path.open() as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if name is not None:
                    lengths.append(len("".join(seq)))
                name = line[1:]
                headers.append(name)
                seq = []
            else:
                if not re.fullmatch(r"[ACGTRYSWKMBDHVNacgtryswkmbdhvn.-]+", line):
                    fail(f"Invalid FASTA characters in {path}")
                seq.append(line)
    if name is not None:
        lengths.append(len("".join(seq)))
    return headers, lengths


def validate_fasta() -> None:
    if not REF_FASTA.exists():
        fail(f"Missing reference FASTA: {REF_FASTA}")
    headers, lengths = read_fasta(REF_FASTA)
    if not headers:
        fail("Reference FASTA contains no records")
    ids = [h.split()[0] for h in headers]
    if len(ids) != len(set(ids)):
        fail("Reference FASTA contains duplicate record IDs")
    missing_group = [h for h in headers if "ddns_group=" not in h]
    if missing_group:
        fail(f"{len(missing_group)} reference headers lack ddns_group metadata")
    if min(lengths) < 500:
        fail("Reference FASTA contains unexpectedly short records")
    groups = set()
    for header in headers:
        match = re.search(r"(?:^|\s)ddns_group=([^\s]+)", header)
        if match:
            groups.add(match.group(1))
    expected = {"WPV1", "WPV2", "WPV3", "Sabin1-related", "Sabin2-related", "Sabin3-related", "NonPolioEV"}
    missing = expected - groups
    if missing:
        fail(f"Reference FASTA missing expected groups: {sorted(missing)}")
    print(f"Reference FASTA OK: {len(headers)} records; groups={len(groups)}")


def validate_primers() -> None:
    if not PRIMERS_XLSX.exists():
        fail(f"Missing primer workbook: {PRIMERS_XLSX}")
    try:
        import openpyxl
    except ImportError as exc:
        fail(f"openpyxl is required for primer validation: {exc}")
    wb = openpyxl.load_workbook(PRIMERS_XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        fail("Primer workbook is empty")
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    needed = {"Name", "Orientation", "Sequence"}
    if not needed.issubset(set(header)):
        fail(f"Primer workbook missing columns: {sorted(needed - set(header))}")
    idx = {name: header.index(name) for name in needed}
    per_bc = {}
    for row in rows[1:]:
        name = row[idx["Name"]]
        orientation = row[idx["Orientation"]]
        seq = row[idx["Sequence"]]
        if not name or not orientation or not seq:
            continue
        match = re.search(r"__(BC\d+)$", str(name))
        if not match:
            continue
        bc = match.group(1)
        per_bc.setdefault(bc, set()).add(str(orientation).strip().lower())
        if not re.fullmatch(r"[ACGTRYSWKMBDHVNacgtryswkmbdhvn]+", str(seq).strip()):
            fail(f"Invalid primer sequence for {name}")
    bad = [bc for bc, orientations in per_bc.items() if orientations != {"forward", "reverse"}]
    if bad:
        fail(f"Barcodes without one forward and one reverse primer: {bad[:10]}")
    print(f"Primer workbook OK: {len(per_bc)} barcodes")


def validate_examples() -> None:
    if not BARCODES_CSV.exists():
        fail(f"Missing example barcode map: {BARCODES_CSV}")
    text = BARCODES_CSV.read_text().strip().splitlines()
    if not text or text[0].strip() != "barcode,sample":
        fail("Example barcode map must start with barcode,sample")
    print("Example barcode map OK")


def main() -> int:
    validate_fasta()
    validate_primers()
    validate_examples()
    return 0


if __name__ == "__main__":
    sys.exit(main())
